#!/usr/bin/python3

import openpyxl
from openpyxl.styles import PatternFill
import datetime


class XlsxTreater:
    def __init__(self, file_to_open, file_to_save):
        self.file_to_open = file_to_open
        self.file_to_save = file_to_save
        workbook = openpyxl.load_workbook(file_to_open)
        sheet = workbook.active
        self.workbook = workbook
        self.sheet = sheet
        self.colors = {'red': '00FF0000', 'green': '0000FF00', 'yellow': '00FFFF00', 'white': '00FFFFFF'}
        self.condition_checker()
        self.workbook.save(file_to_save)

    def condition1(self, row, column=2):
        date = self.sheet.cell(row=row, column=column)
        if date.value < datetime.datetime.today():
            return True
        else:
            return False

    def condition2(self, row, column=3):
        date = self.sheet.cell(row=row, column=column)
        if date.value == 'not done':
            return True
        else:
            return False

    def condition_checker(self, column_to_color=4):
        for rowNum in range(2, self.sheet.max_row):
            if self.condition1(row=rowNum) and self.condition2(row=rowNum):
                selected_color=self.colors['red']
            elif not self.condition1(row=rowNum) and self.condition2(row=rowNum):
                selected_color=self.colors['yellow']
            elif not self.condition2(row=rowNum):
                selected_color=self.colors['green']
            else:
                selected_color=self.colors['white']
            cell_with_color = self.sheet.cell(row=rowNum, column=column_to_color)
            cell_with_color.fill = PatternFill("solid", fgColor=selected_color)


if __name__ == '__main__':
    XlsxTreater('example.xlsx', 'example1.xlsx')
